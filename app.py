from flask import Flask, request, render_template, flash, redirect, url_for, jsonify, make_response, send_from_directory, abort, send_file
from flask.ext.bootstrap import Bootstrap
from flask.ext.mail import Mail, Message
from flask.ext.security import Security, SQLAlchemyUserDatastore, registerable, current_user, login_required, roles_accepted
from flask.ext.sqlalchemy import SQLAlchemy
from flask.ext.storage import get_default_storage_class
from flask.ext.uploads import init, save, Upload

import stripe
import os

from collections import defaultdict
from datetime import datetime

from geopy import geocoders
from openpyxl import load_workbook

g = geocoders.Google()

CONVENIENCE_FEE = 500

STRIPE_KEYS = {
	#"secret_key": "sk_076KrUmlyQWYeTysIbXuapSy7AWcX",
	"secret_key": "sk_076KMBjuzftnlqaTOHhsMfRRk8qhZ",
	#"publishable_key": "pk_076KVxlFN6XEzWvaHaTfES34ccZ7E",
	"publishable_key": "pk_076K5dtDYTT6lRCvpC6HrIWbSeY3a"
}

stripe.api_key = STRIPE_KEYS["secret_key"]

app = Flask(__name__)
app.config.from_pyfile("config.py")

mail = Mail(app)
Bootstrap(app)

from models import Guardian, Role, Player, AgeGroup, Park, ParkStreet, db

db.app = app
db.init_app(app)

user_datastore = SQLAlchemyUserDatastore(db, Guardian, Role)
security = Security(app, user_datastore)

storage = get_default_storage_class(app)
init(db, storage)

def clean_phone_num(phone_num):
	return phone_num.replace("-", "").replace("(", "").replace(")", "").replace(" ", "")

@app.route('/')
def home():
	return render_template("index.html")

from forms import RegistrationForm, PlayerRegistrationForm

@security.register_context_processor
def register_context_processor():
	return dict(user=current_user.get_full_name())

@app.route("/signup", methods=["GET", "POST"])
def signup():
	form = RegistrationForm(request.form)
	park_id = None
	
	if request.method == "POST":
		if form.validate():
			addr = " ".join([form.street.data, form.city.data, "ON", form.postal_code.data])
			try:
				place, (lat, lng) = g.geocode(addr)
				street = " ".join(place.split(",")[0].strip().split()[1:])
				park_street = ParkStreet.query.filter_by(name=street).first()

				if park_street:
					park_id = park_street.park_id
			except:
				msg = Message("No Park Determined!", recipients=["richard.drake@nascsoccer.org"])
				msg.body = "Look into %s %s" % (form.first_name.data, form.last_name.data)

				mail.send(msg)
				pass

			registerable.register_user(
				email = form.email.data,
				password = form.password.data,
				first_name = form.first_name.data,
				last_name = form.last_name.data,
				apt = form.apt.data,
				street = form.street.data,
				city = form.city.data,
				province = "ON",
				postal_code = form.postal_code.data.replace(" ", ""),
				verified_addr = False,
				primary_phone = clean_phone_num(form.primary_phone.data),
				secondary_phone = clean_phone_num(form.secondary_phone.data),
				marketing = form.marketing.data,
				park_id = park_id,
				active = True
			)
			db.session.commit()

			return redirect("/login")
		else:
			flash("Please correct the errors below.", "error")

	return render_template("signup.html", form=form)

@app.route("/register", methods=["GET", "POST"])
@login_required
def register():
	form = PlayerRegistrationForm(request.form)

	if request.method == "POST" and form.validate():
		player = Player(
			first_name = form.first_name.data,
			last_name = form.last_name.data,
			date_of_birth = form.date_of_birth.data,
			gender = form.gender.data,
			played_before = form.played_before.data,
			played_years = form.played_years.data,
			played_position = form.played_position.data,
			osa_another_country = form.osa_another_country.data,
			osa_what_year = form.osa_what_year.data,
			osa_what_country = form.osa_what_country.data,
			osa_what_club = form.osa_what_club.data,
			notes = form.notes.data,
			pooling = form.pooling.data,
			allstar = form.allstar.data,
			ec_name = form.ec_name.data,
			ec_num = clean_phone_num(form.ec_num.data),
			age_group_id = form.age_group_id.data,
			guardian_id = current_user.id,
			paid = False,
			active = True
		)

		db.session.add(player)

		db.session.commit()
		
		flash("Registered!")

		return redirect("/success")

	return render_template("register.html", form=form)

@app.route("/checkout", methods=["GET", "POST"])
@login_required
def checkout():
	# All players who have not been paid for already should show up.
	players = [player for player in current_user.players if not player.paid]
	fees = defaultdict(list)

	# Enumerate through all of the players and create descriptions of each
	# of the fees.
	for player in players:
		ag = AgeGroup.query.get(player.age_group_id)
		p = fees[player.get_full_name()]

		p.append(["%s - %s" % (ag.name, ag.sport), ag.basefee])
		p.append(["User Fee<sup>1</sup>", ag.userfee])
		p.append(["Convenience Fee<sup>2</sup>", CONVENIENCE_FEE])
	
	if len(players) > 0 and current_user.customer_id is None or current_user.customer_id == "":
		if current_user.park_id:
			park = Park.query.get(current_user.park_id)
			fees["Other"].append(["Park Fee - %s<sup>3</sup>" % park.name, park.fee])

	# Total everything up.
	total = 0

	for player in fees:
		for (title, fee) in fees[player]:
			total += fee

	if request.method == "POST":
		# Bill them.  Well, at least try to.
		try:
			if not current_user.customer_id:
				customer = stripe.Customer.create(
					email = current_user.email,
					card = request.form["stripeToken"]
				)

				current_user.customer_id = customer.id
				db.session.commit()

			charge = stripe.Charge.create(
				customer = current_user.customer_id,
				amount = total,
				currency = "cad",
				description = "Registration fee"
			)

			# If the charge went through, mark all of the players as paid.
			if charge.paid:
				for player in players:
					player.paid = True
					player.paid_at = datetime.now()
					db.session.commit()

				msg = Message("Your Registration Was Successful", recipients=[current_user.email])
				msg.body = "We have successfully received your payment of $%0.2f.  Please complete the verification step by submitting supporting documentation to the following page:\n\nhttps://register.nascsoccer.org/verify" % (charge.amount / 100)

				mail.send(msg)

				# Also get verification.
				return redirect("/verify")
		except stripe.CardError as e:
			flash("%s.  Please try again." % e.message, "error")
	
	# If they haven't paid yet, ask them to.
	return render_template("checkout.html", total=total, fees=fees, key=STRIPE_KEYS["publishable_key"])

@app.route("/success")
@login_required
def success():
	return render_template("success.html")

@app.route("/verify")
@login_required
def verify():
	players_verified = [p.verified_dob for p in current_user.players]
	done = current_user.verified_addr and False not in players_verified

	return render_template("verify.html", done=done)

@app.route("/upload", methods=["POST"])
@login_required
def upload():
	if request.method == "POST":
		photo = request.files["photo"]

		# Only save images, ignore any other uploads.
		if "image" in photo.mimetype or "application/pdf" in photo.mimetype:
			save(photo)

			last_inserted_id = Upload.query.all()[-1].id

			player_id = request.form.get("player_id", None)
			player_ids = [player.id for player in current_user.players]

			if player_id and int(player_id) in player_ids:
				Player.query.get(player_id).verified_dob_doc = last_inserted_id
			elif not player_id:
				current_user.verified_addr_doc = last_inserted_id

			msg = Message("Registration Needs Verifying", recipients=["richard.drake@nascsoccer.org"])
			msg.body = "Guardian %d, submitted verification!" % current_user.id

			mail.send(msg)

			# If none of the above are executed, chances are they're trying to
			# change the documentation for a player they haven't registered.

			db.session.commit()
		else:
			flash("Did not receive an image file (PNG, JPEG, GIF, TIFF, etc) or PDF.  Please try again.")

	return redirect("/verify")

# For this to work, need to add role and user to role.
"""
from app import Guardian, Role, db, user_datastore
user = Guardian.query.filter_by(email='rdrake@rdrake.org').first()
user_datastore.create_role(name="admin", description="Administrator, can do anything")
user_datastore.create_role(name="moderator", description="Moderates proof of address/dob/etc.")
db.session.commit()
user_datastore.add_role_to_user(user, "admin")
db.session.commit()
"""
@app.route("/moderate")
@login_required
@roles_accepted("admin", "moderator")
def moderate():
	# Want all addresses that are unverified but have documentation.
	addresses = Guardian.query.filter_by(verified_addr=False).filter(Guardian.verified_addr_doc != None)
	# Players too.
	players = Player.query.filter(Player.verified_dob != True).filter(Player.verified_dob_doc != None)

	return render_template("moderate.html", addresses=addresses, players=players)

@app.route("/waiver/<int:id>")
@login_required
@roles_accepted("admin", "moderator")
def view_waiver(id):
	player = Player.query.get_or_404(id)
	guardian = player.guardian
	ag = AgeGroup.query.get_or_404(player.age_group_id)

	if guardian.park_id is None:
		abort(404)
	
	park = Park.query.get_or_404(guardian.park_id)

	wb = load_workbook(os.path.join(os.path.dirname(__file__), "templates", "waiverformtemplate.xlsx"))
	ws = wb.get_active_sheet()

	ws.cell("E11").value = park.name
	ws.cell("A13").value = player.last_name
	ws.cell("D13").value = player.first_name
	ws.cell("A16").value = player.date_of_birth.strftime("%Y-%m-%d")
	ws.cell("A19").value = guardian.street
	ws.cell("E19").value = guardian.apt
	ws.cell("A22").value = guardian.city
	ws.cell("C22").value = guardian.province
	ws.cell("E22").value = guardian.postal_code
	ws.cell("A25").value = guardian.email
	ws.cell("A30").value = guardian.primary_phone
	ws.cell("D30").value = guardian.secondary_phone
	ws.cell("A35").value = player.ec_name
	ws.cell("D35").value = player.ec_num
	ws.cell("A39").value = player.notes
	ws.cell("A43").value = " ".join([guardian.first_name, guardian.last_name])
	ws.cell("G43").value = player.paid_at.strftime("%Y-%m-%d")
	ws.cell("I12").value = ag.name
	ws.cell("I13").value = "Yes" if len(player.pooling) > 0 else "No"
	ws.cell("I14").value = player.pooling
	ws.cell("I17").value = player.paid_at.strftime("%Y-%m-%d %H:%M:%S")
	ws.cell("I19").value = ag.basefee / 100
	ws.cell("I20").value = ag.userfee / 100
	ws.cell("I21").value = park.fee / 100
	ws.cell("I22").value = 5
	ws.cell("I23").value = (ag.basefee + ag.userfee + park.fee + 500) / 100
	ws.cell("I26").value = "Yes" if player.played_before else "No"
	ws.cell("I29").value = player.played_years
	ws.cell("I32").value = player.played_position

	filename = os.path.join(app.config["WAIVERS_FOLDER"], "%d.xlsx" % id)
	wb.save(filename)
	
	try:
		return send_file(filename)
	except Exception as e:
		abort(404)

@app.route("/api/upload/<int:id>")
@login_required
@roles_accepted("admin", "moderator")
def view_upload(id):
	upload = Upload.query.get_or_404(id)
	return send_from_directory(app.config["UPLOADS_FOLDER"], upload.name)

@app.route("/api/verify/player/<int:id>", methods=["POST"])
@login_required
@roles_accepted("admin", "moderator")
def verify_player(id):
	player = Player.query.get_or_404(id)
	player.verified_dob = True
	player.confirmed_at = datetime.now()
	db.session.commit()

	flash("Successfully verified date of birth for %s, %s" % (player.last_name, player.first_name))

	return redirect("/moderate")

@app.route("/api/verify/address/<int:id>", methods=["POST"])
@login_required
@roles_accepted("admin", "moderator")
def verify_address(id):
	guardian = Guardian.query.get_or_404(id)
	guardian.verified_addr = True
	guardian.confirmed_at = datetime.now()
	db.session.commit()

	flash("Successfully verified address for %s" % guardian.get_full_name())

	return redirect("/moderate")

@app.route("/api/programs")
@login_required
def programs():
	gender = request.args["gender"]
	date_of_birth = request.args["date_of_birth"]

	programs = AgeGroup.query.filter(
		date_of_birth <= AgeGroup.end,
		date_of_birth >= AgeGroup.start,
		AgeGroup.gender.like("%%%s%%" % gender)
	).all()

	return jsonify(programs=[{
		"id": program.id,
		"name": program.name,
		"basefee": program.basefee,
		"userfee": program.userfee,
		"sport": program.sport,
		"text": "%s - %s" % (program.name, program.sport)
	} for program in programs])

if __name__ == "__main__":
	app.run(debug=True, port=9090)

	if app.config['DEBUG']:
		from werkzeug import SharedDataMiddleware
		import os
		app.wsgi_app = SharedDataMiddleware(app.wsgi_app, {
			"/": os.path.join(os.path.dirname(__file__), "static")
		})
